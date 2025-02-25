import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import ftplib
import math
import warnings
import requests

def loadData(csvtoggle, csvpath, useftp, ftpserver, ftppath):
    df = pd.DataFrame()
    root_dirnames = []
    if useftp == "true":
        ftpserver.cwd("Minecraft")
        with open("../data/usercache.json", "wb") as file:
            ftpserver.retrbinary(f"RETR usercache.json", file.write)
        names = pd.DataFrame(json.load(open("../data/usercache.json", "r")))
        ftpserver.cwd("../")

        # Get directories
        root_dirnames = ftpserver.nlst(ftppath)
        ftpserver.cwd(ftppath)
        
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            # Go to the subfolder
            ftpserver.cwd(dirname.split("/")[-1])
            filenames = ftpserver.nlst()
            
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = f"temp_{filename}"
                with open(local_file, "wb") as file:
                    ftpserver.retrbinary(f"RETR {filename}", file.write)
                
                with open(local_file, "r") as file:
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                
                os.remove(local_file)
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                
            ftpserver.cwd("../")  # Move back to the parent directory
    else:
        names_file = open('../../data/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        i = -1
        path = '../../data/world/cobblemonplayerdata'
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open(path + '/' + root_dirnames[i] + '/' + filename)
                data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df


def most_pokemons_leaderboard(df, config, type):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    if type == "standard":
        sheet_name = "leaderboard2"
    elif type == "shiny":
        sheet_name = "leaderboard3"
    elif type == "legendary":
        sheet_name = "leaderboard4"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['ExcelRows'])
    ExcelCols = int(config['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['Subtitle'])
    wb.save(file_path)


def most_pokemons_leaderboard_html(df, config, type):
    num_players = len(df)

    #date based on Paris timezone
    now = datetime.datetime.now(tz=datetime.timezone(datetime.timedelta(hours=1)))
    last_updated = now.strftime(config['LastUpdated'])
    subtitle = config['Subtitle']

    leaderboard_class = type.lower()

    html = "<html>\n<head>\n"
    html += f"<meta charset='utf-8'>\n<title>Leaderboard - {type.capitalize()}</title>\n"

    html += '<link rel="stylesheet" type="text/css" href="leaderboard_styles.css">\n'

    html += "</head>\n<body>\n"
    html += '<div class="container">\n'
    html += f"<h2>Leaderboard - {type.capitalize()}</h2>\n"
    html += f'<table class="{leaderboard_class}">\n'

    # Add table headers
    html += "  <tr>\n"
    html += "    <th>Rang</th>\n"
    html += "    <th id='header-player'>Joueur</th>\n"
    html += "    <th>Capturés</th>\n"
    html += "  </tr>\n"

    for i, (index, row) in enumerate(df.iterrows(), start=1):
        html += "  <tr>\n"
        html += f"    <td>{i}</td>\n"
        html += (f"    <td>"
                    f"    <div class='player-container'>"
        f" <img src='https://minotar.net/helm/{index}/32.png' alt='{index}' class='player-face'>"
        f"<span class='player-text'>{index}</span></div>"
                 f"</td>\n")
        html += f"    <td>{row.iloc[0]}</td>\n"
        html += "  </tr>\n"

    html += "</table>\n"

    html += f'<p class="footer">{last_updated}</p>\n'
    html += "</div>\n"
    html += "</body>\n</html>"

    output_file_mapping = {
        "standard": "./html/output_standard.html",
        "shiny": "./html/output_shiny.html",
        "legendary": "./html/output_legendary.html"
    }
    output_file = output_file_mapping.get(type, "output.html")

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Leaderboard saved to {output_file}")


# Read config
config = configparser.ConfigParser()
config.read('cobblemon_config.ini', encoding='utf8')

# Connect to FTP if activated
ftp_server = None

if config['FTP']['UseFTP'] == "true":
    ftp_server = ftplib.FTP(config['FTP']['Host'], open("../username.txt", "r").read(), open("../password.txt", "r").read())
    ftp_server.encoding = "utf-8"

# Load the data
if config['GLOBALMATRIX']['UseCSV'] == "false":
    df = loadData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['FTP']['UseFTP'], ftp_server, config['FTP']['Path'])
else:
    df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if config['FTP']['UseFTP'] == "true":
    ftp_server.quit()

# Prepare the counting DF
count_df = df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

# Leaderboard feature
if config['LEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard_html(player_sum, config['LEADERBOARD'], "standard")

# Shiny leaderboard feature
if config['SHINYLEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame(((df == "True") | (df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['SHINYLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard_html(player_sum, config['SHINYLEADERBOARD'], "shiny")

# Legendary leaderboard feature
if config['LEGLEADERBOARD']['Enable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    #leg_count_df.to_csv("temp.csv")
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEGLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard_html(player_sum, config['LEGLEADERBOARD'], "legendary")

print("Done!")